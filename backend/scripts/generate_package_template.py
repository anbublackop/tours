"""
Generate an Excel template for adding tour packages and destinations.

Usage:
    pip install openpyxl
    python scripts/generate_package_template.py

Output: package_template.xlsx (in the backend/ directory)

Sheet layout
────────────
  1. Packages       — core flat fields (one row per package)
  2. Itinerary      — day-wise plan with optional photo URL
  3. Hotels         — accommodation options per package
  4. Transport      — transport options per package
  5. Addons         — optional add-ons per package
  6. Lists          — inclusions / exclusions / booking_rules / travel_rules
                      (one item per row — much easier than pipe-separated cells)
  7. Destinations   — country/destination records (slug, name, image, banner)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark navy
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="E8F4FD")   # light blue
NOTE_FILL    = PatternFill("solid", fgColor="F2F2F2")   # light grey
NOTE_FONT    = Font(italic=True, size=9, color="555555")

# Sheet-tab colours
TAB_BLUE   = "1F4E79"
TAB_GREEN  = "375623"
TAB_ORANGE = "C55A11"
TAB_PURPLE = "7030A0"
TAB_TEAL   = "006060"
TAB_RED    = "C00000"
TAB_BROWN  = "7B3F00"

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _note_row(ws, row, notes):
    for col, note in enumerate(notes, 1):
        cell = ws.cell(row=row, column=col, value=note)
        cell.font      = NOTE_FONT
        cell.fill      = NOTE_FILL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _example_row(ws, row, data, col_count):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = EXAMPLE_FILL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _blank_rows(ws, start, end, col_count):
    for row in range(start, end + 1):
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _auto_width(ws, min_w=14, max_w=52):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_w), max_w
        )


def _legend(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="1F4E79", size=10)


# ── Sheet 1 — Packages ───────────────────────────────────────────────────────

def build_packages_sheet(wb):
    ws = wb.active
    ws.title = "Packages"
    ws.sheet_properties.tabColor = TAB_BLUE

    headers = [
        ("title *",         "Full package name — must be UNIQUE and must match exactly in all other sheets"),
        ("country *",       "india | nepal | south-korea | thailand | china | sri-lanka"),
        ("category *",      "Wildlife Safari | Religious & Spiritual | Adventure | Heritage & Culture | Beach & Coastal | Hill Stations"),
        ("state",           "State / region, e.g. 'Rajasthan'  (optional)"),
        ("location",        "Main city/area, e.g. 'Jaipur, Jodhpur, Jaisalmer'"),
        ("description *",   "1-3 sentence overview shown on the package detail page"),
        ("price *",         "Base price per person in INR (integer), e.g. 45000"),
        ("original_price",  "MRP / strike-through price in INR — leave blank if no discount"),
        ("duration *",      "Display string, e.g. '7D/6N'"),
        ("duration_days *", "Total days as integer, e.g. 7"),
        ("image",           "Cover image URL (optional — a placeholder is used if blank)"),
        ("rating",          "Display rating, e.g. '4.8'  (optional, defaults to 4.5)"),
        ("reviews_count",   "Number of reviews to display, e.g. 128  (optional)"),
        ("highlights",      "Pipe-separated highlights shown on the listing card, e.g. 'Amber Fort|Camel Safari|Desert Camp'"),
        ("available_dates", "Free-text season note, e.g. 'Oct – Mar'  (optional)"),
        ("max_group_size",  "Max travellers per departure, e.g. 20  (optional)"),
    ]

    for col, (h, _) in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))

    _note_row(ws, 2, [n for _, n in headers])
    ws.row_dimensions[2].height = 55

    example = [
        "Rajasthan Royal Heritage Tour",
        "india",
        "Heritage & Culture",
        "Rajasthan",
        "Jaipur, Jodhpur, Jaisalmer",
        "An unforgettable journey through the royal palaces and golden deserts of Rajasthan.",
        45000, 55000, "7D/6N", 7,
        "https://images.unsplash.com/photo-1524492412937-b28074a5d7da?w=800",
        "4.8", 128,
        "Amber Fort|Camel Safari|Desert Camp|City Palace|Mehrangarh Fort",
        "Oct – Mar", 20,
    ]
    _example_row(ws, 3, example, len(headers))
    _blank_rows(ws, 4, 14, len(headers))

    ws.freeze_panes = "A3"
    _auto_width(ws)
    _legend(ws, 16,
        "* = required   |   Row 3 is an example — replace or delete it   |   "
        "Inclusions / Exclusions / Rules → use the 'Lists' sheet"
    )


# ── Sheet 2 — Itinerary ──────────────────────────────────────────────────────

def build_itinerary_sheet(wb):
    ws = wb.create_sheet("Itinerary")
    ws.sheet_properties.tabColor = TAB_GREEN

    headers = [
        "package_title *",
        "day *",
        "title *",
        "description *",
        "meals",
        "overnight",
        "image_url",
    ]
    notes = [
        "Must match title in Packages sheet exactly",
        "Day number as integer, e.g. 1",
        "Short day title, e.g. 'Arrive in Jaipur – The Pink City'",
        "Full description of the day's activities and highlights",
        "Pipe-separated meals included, e.g. 'Breakfast|Dinner'  (optional)",
        "City/place where guests stay that night, e.g. 'Jaipur'  (optional)",
        "URL of a photo representing this day (optional — shown in itinerary section)",
    ]
    examples = [
        ["Rajasthan Royal Heritage Tour", 1,
         "Arrive in Jaipur – The Pink City",
         "Arrive at Jaipur airport, transfer to hotel. Evening stroll to Hawa Mahal and local bazaars.",
         "Dinner", "Jaipur",
         "https://images.unsplash.com/photo-1599661046289-e31897846e41?w=600"],
        ["Rajasthan Royal Heritage Tour", 2,
         "Amber Fort & City Palace",
         "Morning elephant ride up to Amber Fort. Afternoon explore the City Palace and Jantar Mantar observatory.",
         "Breakfast|Dinner", "Jaipur",
         "https://images.unsplash.com/photo-1524492412937-b28074a5d7da?w=600"],
        ["Rajasthan Royal Heritage Tour", 3,
         "Jaipur → Jodhpur – The Blue City",
         "Scenic drive to Jodhpur. Check-in and sunset visit to Mehrangarh Fort with panoramic city views.",
         "Breakfast", "Jodhpur",
         ""],
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))
    _note_row(ws, 2, notes)
    ws.row_dimensions[2].height = 45

    for r, row_data in enumerate(examples, 3):
        _example_row(ws, r, row_data, len(headers))

    _blank_rows(ws, 6, 40, len(headers))
    ws.freeze_panes = "A3"
    _auto_width(ws)
    _legend(ws, 42, "Leave image_url blank if no photo is available for that day.")


# ── Sheet 3 — Hotels ─────────────────────────────────────────────────────────

def build_hotels_sheet(wb):
    ws = wb.create_sheet("Hotels")
    ws.sheet_properties.tabColor = TAB_ORANGE

    headers = ["package_title *", "hotel_id *", "name *", "category *",
               "pricePerNight *", "rating", "amenities"]
    notes = [
        "Must match title in Packages sheet exactly",
        "Unique short ID per package, e.g. 'h1' — used by checkout to select hotel",
        "Hotel display name, e.g. 'Rambagh Palace'",
        "budget | standard | luxury | ultra-luxury",
        "Extra price per person per night in INR (integer)",
        "Display rating string, e.g. '4.5'  (optional)",
        "Pipe-separated amenities, e.g. 'Pool|WiFi|Spa|Restaurant'  (optional)",
    ]
    examples = [
        ["Rajasthan Royal Heritage Tour", "h1", "Budget Inn Jaipur",  "budget",       1500, "3.8", "WiFi|AC"],
        ["Rajasthan Royal Heritage Tour", "h2", "Clarks Amer Jaipur", "standard",     3500, "4.2", "WiFi|Pool|Restaurant"],
        ["Rajasthan Royal Heritage Tour", "h3", "Rambagh Palace",     "luxury",       9500, "4.8", "Pool|Spa|WiFi|Fine Dining|Butler"],
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))
    _note_row(ws, 2, notes)
    ws.row_dimensions[2].height = 40
    for r, row_data in enumerate(examples, 3):
        _example_row(ws, r, row_data, len(headers))
    _blank_rows(ws, 6, 30, len(headers))
    ws.freeze_panes = "A3"
    _auto_width(ws)


# ── Sheet 4 — Transport ──────────────────────────────────────────────────────

def build_transport_sheet(wb):
    ws = wb.create_sheet("Transport")
    ws.sheet_properties.tabColor = TAB_ORANGE

    headers = ["package_title *", "transport_id *", "type *", "description *", "price *"]
    notes = [
        "Must match title in Packages sheet exactly",
        "Unique short ID, e.g. 't1' — used by checkout to select transport",
        "Display label, e.g. 'Shared Bus' or 'Private SUV'",
        "Short description shown to the traveller",
        "Extra price per person in INR (integer) — use 0 if included in base price",
    ]
    examples = [
        ["Rajasthan Royal Heritage Tour", "t1", "Shared Bus",    "AC Volvo coach, shared with group (max 20)",           0],
        ["Rajasthan Royal Heritage Tour", "t2", "Private SUV",   "Dedicated Innova Crysta for your group (up to 6 pax)", 4000],
        ["Rajasthan Royal Heritage Tour", "t3", "Private Tempo", "Tempo Traveller for groups up to 12",                  2500],
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))
    _note_row(ws, 2, notes)
    ws.row_dimensions[2].height = 40
    for r, row_data in enumerate(examples, 3):
        _example_row(ws, r, row_data, len(headers))
    _blank_rows(ws, 6, 20, len(headers))
    ws.freeze_panes = "A3"
    _auto_width(ws)


# ── Sheet 5 — Addons ─────────────────────────────────────────────────────────

def build_addons_sheet(wb):
    ws = wb.create_sheet("Addons")
    ws.sheet_properties.tabColor = TAB_ORANGE

    headers = ["package_title *", "addon_id *", "name *", "description *", "price *", "icon"]
    notes = [
        "Must match title in Packages sheet exactly",
        "Unique short ID, e.g. 'a1'",
        "Add-on display name, e.g. 'Camel Safari'",
        "Short description shown to the traveller",
        "Extra price per person in INR (integer)",
        "Emoji icon displayed next to the add-on, e.g. 🐪  (optional)",
    ]
    examples = [
        ["Rajasthan Royal Heritage Tour", "a1", "Camel Safari",         "1-hour sunset camel ride in the Thar Desert",           800,  "🐪"],
        ["Rajasthan Royal Heritage Tour", "a2", "Desert Camp Stay",     "Overnight stay in a luxury desert camp with folk music", 3500, "⛺"],
        ["Rajasthan Royal Heritage Tour", "a3", "Hot Air Balloon Ride", "Sunrise balloon flight over Jaipur",                    6500, "🎈"],
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))
    _note_row(ws, 2, notes)
    ws.row_dimensions[2].height = 40
    for r, row_data in enumerate(examples, 3):
        _example_row(ws, r, row_data, len(headers))
    _blank_rows(ws, 6, 20, len(headers))
    ws.freeze_panes = "A3"
    _auto_width(ws)


# ── Sheet 6 — Lists (inclusions / exclusions / rules) ────────────────────────

def build_lists_sheet(wb):
    """
    One row per list item.  list_type must be one of:
      inclusions | exclusions | booking_rules | travel_rules
    """
    ws = wb.create_sheet("Lists")
    ws.sheet_properties.tabColor = TAB_PURPLE

    headers = ["package_title *", "list_type *", "item *"]
    notes = [
        "Must match title in Packages sheet exactly",
        "inclusions  |  exclusions  |  booking_rules  |  travel_rules",
        "One item per row — e.g. 'Daily breakfast included' or 'Airfare not included'",
    ]
    examples = [
        # Inclusions
        ["Rajasthan Royal Heritage Tour", "inclusions", "Daily breakfast at the hotel"],
        ["Rajasthan Royal Heritage Tour", "inclusions", "AC transport throughout the tour"],
        ["Rajasthan Royal Heritage Tour", "inclusions", "English-speaking licensed guide"],
        ["Rajasthan Royal Heritage Tour", "inclusions", "All monument entry fees"],
        ["Rajasthan Royal Heritage Tour", "inclusions", "Welcome drink and traditional Rajasthani dinner on Day 1"],
        # Exclusions
        ["Rajasthan Royal Heritage Tour", "exclusions", "Airfare / train fare to Jaipur and back"],
        ["Rajasthan Royal Heritage Tour", "exclusions", "Lunch, dinner (except Day 1 welcome dinner)"],
        ["Rajasthan Royal Heritage Tour", "exclusions", "Personal expenses and tips"],
        ["Rajasthan Royal Heritage Tour", "exclusions", "Travel insurance"],
        # Booking rules
        ["Rajasthan Royal Heritage Tour", "booking_rules", "50% advance payment required at booking"],
        ["Rajasthan Royal Heritage Tour", "booking_rules", "Remaining balance due 15 days before travel date"],
        ["Rajasthan Royal Heritage Tour", "booking_rules", "No changes or cancellations within 7 days of departure"],
        ["Rajasthan Royal Heritage Tour", "booking_rules", "Full refund if cancelled more than 30 days before departure"],
        # Travel rules
        ["Rajasthan Royal Heritage Tour", "travel_rules", "Carry a valid government-issued photo ID"],
        ["Rajasthan Royal Heritage Tour", "travel_rules", "Comfortable walking shoes are strongly recommended"],
        ["Rajasthan Royal Heritage Tour", "travel_rules", "Dress modestly when visiting temples and religious sites"],
        ["Rajasthan Royal Heritage Tour", "travel_rules", "No smoking inside vehicles or heritage buildings"],
    ]

    # Color-code the list_type column with fill per type
    TYPE_FILLS = {
        "inclusions":    PatternFill("solid", fgColor="E2EFDA"),  # green
        "exclusions":    PatternFill("solid", fgColor="FCE4D6"),  # red-orange
        "booking_rules": PatternFill("solid", fgColor="DDEBF7"),  # blue
        "travel_rules":  PatternFill("solid", fgColor="FFF2CC"),  # yellow
    }

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))
    _note_row(ws, 2, notes)
    ws.row_dimensions[2].height = 35

    for r, row_data in enumerate(examples, 3):
        list_type = row_data[1]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill   = TYPE_FILLS.get(list_type, EXAMPLE_FILL)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _blank_rows(ws, len(examples) + 3, len(examples) + 60, len(headers))
    ws.freeze_panes = "A3"
    _auto_width(ws)

    # Colour key
    start = len(examples) + 5
    ws.cell(row=start, column=1, value="Colour key:").font = Font(bold=True, size=10)
    for i, (lt, fill) in enumerate(TYPE_FILLS.items(), 1):
        c = ws.cell(row=start + i, column=1, value=lt)
        c.fill   = fill
        c.border = THIN_BORDER
        c.font   = Font(size=9)


# ── Sheet 7 — Destinations ───────────────────────────────────────────────────

def build_destinations_sheet(wb):
    """
    Each row creates or updates a destination record in the destinations table.
    Destinations are the country-level landing pages (India, Nepal, etc.).
    """
    ws = wb.create_sheet("Destinations")
    ws.sheet_properties.tabColor = TAB_TEAL

    headers = [
        "slug *",
        "name *",
        "description *",
        "image_url",
        "banner_url",
        "is_featured",
    ]
    notes = [
        "URL-safe identifier used in routes, e.g. 'india' or 'south-korea' — must be unique",
        "Display name shown on the site, e.g. 'India'",
        "2-4 sentence description shown on the destination landing page",
        "Card thumbnail image URL (shown in the destinations grid on the home page)",
        "Full-width banner image URL (shown at the top of the destination packages page)",
        "TRUE or FALSE — whether this destination appears in the featured destinations grid",
    ]
    examples = [
        ["india",
         "India",
         "A land of royal palaces, ancient temples, vibrant festivals, and breathtaking landscapes — from the Himalayas to the backwaters of Kerala.",
         "https://images.unsplash.com/photo-1524492412937-b28074a5d7da?w=800&q=85",
         "https://images.unsplash.com/photo-1524492412937-b28074a5d7da?w=1400&q=85",
         "TRUE"],
        ["nepal",
         "Nepal",
         "Home to eight of the world's ten highest peaks, Nepal offers epic Himalayan treks, ancient temples, and vibrant cultural experiences.",
         "https://images.unsplash.com/photo-1544735716-392fe2489ffa?w=800&q=85",
         "https://images.unsplash.com/photo-1544735716-392fe2489ffa?w=1400&q=85",
         "TRUE"],
        ["south-korea",
         "South Korea",
         "A seamless blend of ultra-modern cities and ancient palaces, K-culture, neon-lit street food alleys, and serene mountain temples.",
         "https://images.unsplash.com/photo-1517154421773-0529f29ea451?w=800&q=85",
         "https://images.unsplash.com/photo-1517154421773-0529f29ea451?w=1400&q=85",
         "TRUE"],
        ["thailand",
         "Thailand",
         "Golden temples, pristine island beaches, aromatic street food, and warm hospitality — Thailand is Asia's most beloved travel destination.",
         "https://images.unsplash.com/photo-1528360983277-13d401cdc186?w=800&q=85",
         "https://images.unsplash.com/photo-1528360983277-13d401cdc186?w=1400&q=85",
         "TRUE"],
        ["china",
         "China",
         "Ancient dynasties meet futuristic skylines. Walk the Great Wall, discover the Terracotta Army, cruise the Yangtze, and explore vibrant modern cities.",
         "https://images.unsplash.com/photo-1547981609-4b6bfe67ca0b?w=800&q=85",
         "https://images.unsplash.com/photo-1547981609-4b6bfe67ca0b?w=1400&q=85",
         "TRUE"],
        ["sri-lanka",
         "Sri Lanka",
         "The teardrop isle packs ancient kingdoms, misty tea highlands, golden beaches, diverse wildlife, and some of Asia's finest surf into one compact paradise.",
         "https://images.unsplash.com/photo-1581888227599-779811939961?w=800&q=85",
         "https://images.unsplash.com/photo-1581888227599-779811939961?w=1400&q=85",
         "TRUE"],
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header(ws, 1, len(headers))
    _note_row(ws, 2, notes)
    ws.row_dimensions[2].height = 50

    for r, row_data in enumerate(examples, 3):
        _example_row(ws, r, row_data, len(headers))

    _blank_rows(ws, len(examples) + 3, len(examples) + 15, len(headers))
    ws.freeze_panes = "A3"
    _auto_width(ws)
    _legend(ws, len(examples) + 17,
        "* = required   |   is_featured controls whether the destination card appears on the homepage"
    )


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    build_packages_sheet(wb)
    build_itinerary_sheet(wb)
    build_hotels_sheet(wb)
    build_transport_sheet(wb)
    build_addons_sheet(wb)
    build_lists_sheet(wb)
    build_destinations_sheet(wb)

    out = "package_template.xlsx"
    wb.save(out)
    print(f"✅  Template saved → {out}")
    print()
    print("Sheets:")
    print("  1. Packages     — core package fields (one row per package)")
    print("  2. Itinerary    — day-by-day plan with optional photo URL per day")
    print("  3. Hotels       — accommodation options")
    print("  4. Transport    — transport options")
    print("  5. Addons       — optional extras")
    print("  6. Lists        — inclusions / exclusions / booking_rules / travel_rules")
    print("                    (one item per row — colour-coded by type)")
    print("  7. Destinations — country landing pages (slug, images, description)")
    print()
    print("Next steps:")
    print("  1. Fill in your data (delete example rows first).")
    print("  2. Dry-run:  python scripts/import_packages.py package_template.xlsx --dry-run")
    print("  3. Import:   python scripts/import_packages.py package_template.xlsx")


if __name__ == "__main__":
    main()
