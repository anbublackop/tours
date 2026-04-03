"""
Import tour packages and destinations from an Excel template or Google Sheets URL.

Usage:
    cd backend/

    # Local file
    python scripts/import_packages.py package_template.xlsx [--dry-run]

    # Google Sheets (sheet must be shared as "Anyone with the link can view")
    python scripts/import_packages.py "https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/..." [--dry-run]

Options:
    --dry-run   Parse and print what would be inserted without touching the DB.
"""

import sys
import io
import os
import re
from datetime import datetime

# ── Make backend package importable ──────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import requests
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"))

from app.db.session import SessionLocal
# Import ALL models so SQLAlchemy can resolve every relationship string
# (e.g. Package.bookings → "Booking") before the mapper is finalized.
from app.models.destination import Destination
from app.models.package_category import PackageCategory
from app.models.package import Package
from app.models.user import User          # noqa: F401
from app.models.booking import Booking    # noqa: F401
from app.models.review import Review      # noqa: F401
from app.models.enquiry import Enquiry    # noqa: F401
try:
    from app.models.refresh_tokens import RefreshToken  # noqa: F401
except Exception:
    pass


# ── Google Sheets / file loader ───────────────────────────────────────────────

_GSHEETS_RE = re.compile(
    r"https://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)"
)

def load_workbook(source: str):
    """
    Load an openpyxl workbook from either:
      • a local file path  (anything that doesn't look like a Google Sheets URL)
      • a Google Sheets URL  (sheet must be shared as "Anyone with the link can view")

    Returns the openpyxl Workbook object.
    """
    m = _GSHEETS_RE.search(source)
    if m:
        spreadsheet_id = m.group(1)
        export_url = (
            f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
            f"/export?format=xlsx"
        )
        print(f"🌐  Downloading Google Sheet (id={spreadsheet_id}) …")
        resp = requests.get(export_url, timeout=30)
        if resp.status_code != 200:
            print(
                f"❌  Failed to download sheet — HTTP {resp.status_code}.\n"
                f"    Make sure the sheet is shared as 'Anyone with the link can view'."
            )
            sys.exit(1)
        return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

    # Local file
    if not os.path.exists(source):
        print(f"❌  File not found: {source}")
        sys.exit(1)
    print(f"📂  Reading {source} …")
    return openpyxl.load_workbook(source, data_only=True)


# ── Cell / sheet helpers ──────────────────────────────────────────────────────

def _str(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None and str(v).strip() else None


def _int(ws, row, col):
    v = ws.cell(row=row, column=col).value
    try:
        return int(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _bool(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return True
    return str(v).strip().upper() not in ("FALSE", "0", "NO", "N")


def _pipe(value):
    if not value:
        return []
    return [item.strip() for item in str(value).split("|") if item.strip()]


def _headers(ws):
    """Return {clean_name: col_index} from row 1."""
    return {
        str(ws.cell(row=1, column=col).value or "").replace(" *", "").strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=col).value
    }


def _col(h_map, *names):
    """Return column index for the first matching header name (or 999 if missing)."""
    for n in names:
        if n in h_map:
            return h_map[n]
    return 999


def _is_metadata(value: str | None) -> bool:
    """
    Return True if a cell value looks like a legend / footer / colour-key row
    rather than real data.  These rows are written by _legend() in the template
    generator into column 1 and get picked up by ws.max_row, causing phantom rows.
    """
    if not value:
        return False
    v = value.strip()
    return (
        v.startswith("*")           # e.g. "* = required | Row 3 is an example…"
        or v.startswith("Colour key")
        or "* = required" in v
        or "Leave " in v and " blank" in v   # e.g. "Leave image_url blank…"
        or len(v) > 120              # no real title / slug is this long
    )


# ── Sheet readers ─────────────────────────────────────────────────────────────

def read_packages(ws):
    h = _headers(ws)
    result = []
    for row in range(3, ws.max_row + 1):
        title = _str(ws, row, _col(h, "title"))
        if not title or _is_metadata(title):
            continue
        # destination_slug is the new column name; fall back to country for old sheets
        destination_slug = _str(ws, row, _col(h, "destination_slug", "country"))
        # category_slug is the new column name; fall back to category string for old sheets
        category_slug    = _str(ws, row, _col(h, "category_slug"))
        category_str     = _str(ws, row, _col(h, "category"))
        result.append({
            "_row":             row,                # actual sheet row for error messages
            "title":            title,
            "destination_slug": destination_slug,
            "country":          destination_slug,   # kept for DB column compat
            "category_slug":    category_slug,
            "category":         category_str,       # legacy string column
            "state":            _str(ws, row, _col(h, "state")),
            "location":         _str(ws, row, _col(h, "location")),
            "description":      _str(ws, row, _col(h, "description")),
            "price":            _int(ws, row, _col(h, "price")) or 0,
            "original_price":   _int(ws, row, _col(h, "original_price")),
            "duration":         _str(ws, row, _col(h, "duration")),
            "duration_days":    _int(ws, row, _col(h, "duration_days")),
            "image":            _str(ws, row, _col(h, "image")),
            "rating":           _str(ws, row, _col(h, "rating")),
            "reviews_count":    _int(ws, row, _col(h, "reviews_count")),
            "highlights":       _pipe(_str(ws, row, _col(h, "highlights"))),
            "available_dates":  _str(ws, row, _col(h, "available_dates")),
            "max_group_size":   _int(ws, row, _col(h, "max_group_size")),
        })
    return result


def read_itinerary(ws):
    """Returns {package_title: [day_dict, ...]} sorted by day number."""
    h = _headers(ws)
    result = {}
    for row in range(3, ws.max_row + 1):
        pkg = _str(ws, row, _col(h, "package_title"))
        if not pkg or _is_metadata(pkg):
            continue
        meals_raw = _str(ws, row, _col(h, "meals"))
        entry = {
            "day":         _int(ws, row, _col(h, "day")) or 0,
            "title":       _str(ws, row, _col(h, "title")) or "",
            "description": _str(ws, row, _col(h, "description")) or "",
            "meals":       _pipe(meals_raw),
            "overnight":   _str(ws, row, _col(h, "overnight")) or "",
            "image_url":   _str(ws, row, _col(h, "image_url")) or "",
        }
        result.setdefault(pkg, []).append(entry)
    for pkg in result:
        result[pkg].sort(key=lambda d: d["day"])
    return result


def read_hotels(ws):
    h = _headers(ws)
    result = {}
    for row in range(3, ws.max_row + 1):
        pkg = _str(ws, row, _col(h, "package_title"))
        if not pkg or _is_metadata(pkg):
            continue
        entry = {
            "id":           _str(ws, row, _col(h, "hotel_id")) or f"h{row}",
            "name":         _str(ws, row, _col(h, "name")) or "",
            "category":     _str(ws, row, _col(h, "category")) or "standard",
            "pricePerNight":_int(ws, row, _col(h, "pricePerNight")) or 0,
            "rating":       _str(ws, row, _col(h, "rating")) or "4.0",
            "amenities":    _pipe(_str(ws, row, _col(h, "amenities"))),
        }
        result.setdefault(pkg, []).append(entry)
    return result


def read_transport(ws):
    h = _headers(ws)
    result = {}
    for row in range(3, ws.max_row + 1):
        pkg = _str(ws, row, _col(h, "package_title"))
        if not pkg or _is_metadata(pkg):
            continue
        entry = {
            "id":          _str(ws, row, _col(h, "transport_id")) or f"t{row}",
            "type":        _str(ws, row, _col(h, "type")) or "",
            "description": _str(ws, row, _col(h, "description")) or "",
            "price":       _int(ws, row, _col(h, "price")) or 0,
        }
        result.setdefault(pkg, []).append(entry)
    return result


def read_addons(ws):
    h = _headers(ws)
    result = {}
    for row in range(3, ws.max_row + 1):
        pkg = _str(ws, row, _col(h, "package_title"))
        if not pkg or _is_metadata(pkg):
            continue
        entry = {
            "id":          _str(ws, row, _col(h, "addon_id")) or f"a{row}",
            "name":        _str(ws, row, _col(h, "name")) or "",
            "description": _str(ws, row, _col(h, "description")) or "",
            "price":       _int(ws, row, _col(h, "price")) or 0,
            "icon":        _str(ws, row, _col(h, "icon")) or "",
        }
        result.setdefault(pkg, []).append(entry)
    return result


def read_lists(ws):
    """
    Returns {package_title: {list_type: [items]}}
    list_type ∈ {inclusions, exclusions, booking_rules, travel_rules}
    """
    h = _headers(ws)
    result = {}
    valid_types = {"inclusions", "exclusions", "booking_rules", "travel_rules"}
    for row in range(3, ws.max_row + 1):
        pkg       = _str(ws, row, _col(h, "package_title"))
        list_type = _str(ws, row, _col(h, "list_type"))
        item      = _str(ws, row, _col(h, "item"))
        if not pkg or not list_type or not item or _is_metadata(pkg):
            continue
        list_type = list_type.strip().lower()
        if list_type not in valid_types:
            print(f"   ⚠️   Row {row} in Lists: unknown list_type '{list_type}' — skipping")
            continue
        result.setdefault(pkg, {}).setdefault(list_type, []).append(item)
    return result


def read_destinations(ws):
    h = _headers(ws)
    result = []
    for row in range(3, ws.max_row + 1):
        slug = _str(ws, row, _col(h, "slug"))
        if not slug or _is_metadata(slug):
            continue
        result.append({
            "_row":        row,
            "slug":        slug,
            "name":        _str(ws, row, _col(h, "name")),
            "description": _str(ws, row, _col(h, "description")),
            "image_url":   _str(ws, row, _col(h, "image_url")),
            "banner_url":  _str(ws, row, _col(h, "banner_url")),
            "is_featured": _bool(ws, row, _col(h, "is_featured")),
        })
    return result


def read_categories(ws):
    h = _headers(ws)
    result = []
    for row in range(3, ws.max_row + 1):
        slug = _str(ws, row, _col(h, "slug"))
        if not slug or _is_metadata(slug):
            continue
        result.append({
            "_row":        row,
            "slug":        slug,
            "name":        _str(ws, row, _col(h, "name")),
            "description": _str(ws, row, _col(h, "description")),
            "icon":        _str(ws, row, _col(h, "icon")),
        })
    return result


# ── Validation ────────────────────────────────────────────────────────────────

def validate_packages(packages):
    errors = []
    for p in packages:
        row = p.get("_row", "?")
        for field in ("title", "description", "duration"):
            if not p.get(field):
                errors.append(f"Packages sheet row {row}: missing required field '{field}'")
        if not (p.get("destination_slug") or p.get("country")):
            errors.append(f"Packages sheet row {row}: missing 'destination_slug' (references Destinations sheet)")
        if not (p.get("category_slug") or p.get("category")):
            errors.append(f"Packages sheet row {row}: missing 'category_slug' (references Categories sheet)")
        if not p.get("price"):
            errors.append(f"Packages sheet row {row}: 'price' must be a positive integer")
    return errors


def validate_destinations(destinations):
    errors = []
    for d in destinations:
        row = d.get("_row", "?")
        for field in ("slug", "name", "description"):
            if not d.get(field):
                errors.append(f"Destinations sheet row {row}: missing required field '{field}'")
    return errors


# ── Import helpers ────────────────────────────────────────────────────────────

def import_packages(packages, itinerary, hotels, transport, addons, lists_data,
                    db, dry_run):
    now = str(datetime.utcnow())
    inserted = updated = 0

    # Build lookup maps for FK resolution
    dest_id_map = {}
    cat_id_map  = {}
    if not dry_run:
        for d in db.query(Destination).all():
            dest_id_map[d.slug] = d.id
        for c in db.query(PackageCategory).all():
            cat_id_map[c.slug] = c.id

    for pkg_data in packages:
        title = pkg_data["title"]
        lists = lists_data.get(title, {})
        country        = pkg_data.get("destination_slug") or pkg_data.get("country") or ""
        destination_id = dest_id_map.get(country)
        category_slug  = pkg_data.get("category_slug")
        category_id    = cat_id_map.get(category_slug) if category_slug else None

        fields = dict(
            name           = title,
            destination_id = destination_id,
            category_id    = category_id,
            country        = country,
            category       = pkg_data.get("category"),
            state          = pkg_data["state"],
            location       = pkg_data["location"],
            description    = pkg_data["description"],
            price          = pkg_data["price"],
            original_price = pkg_data["original_price"],
            duration       = pkg_data["duration"],
            duration_days  = pkg_data["duration_days"],
            image          = pkg_data["image"],
            rating         = pkg_data["rating"] or "4.5",
            reviews_count  = pkg_data["reviews_count"] or 0,
            highlights     = pkg_data["highlights"] or [],
            itinerary      = itinerary.get(title, []),
            hotels         = hotels.get(title, []),
            transport      = transport.get(title, []),
            addons         = addons.get(title, []),
            inclusions     = lists.get("inclusions", []),
            exclusions     = lists.get("exclusions", []),
            booking_rules  = lists.get("booking_rules", []),
            travel_rules   = lists.get("travel_rules", []),
            available_dates= pkg_data["available_dates"],
            max_group_size = pkg_data["max_group_size"],
            updated_at     = now,
        )

        if dry_run:
            itin_days = itinerary.get(title, [])
            days_with_photo = sum(1 for d in itin_days if d.get("image_url"))
            print(f"   ✅  [DRY RUN] '{title}'")
            print(f"       destination={country}  destination_id={destination_id or '(not linked)'}  category_slug={category_slug or '—'}  category_id={category_id or '(not linked)'}  price=₹{fields['price']}")
            print(f"       itinerary: {len(itin_days)} days ({days_with_photo} with photo)")
            print(f"       hotels: {len(hotels.get(title, []))}  transport: {len(transport.get(title, []))}  addons: {len(addons.get(title, []))}")
            print(f"       inclusions: {len(lists.get('inclusions', []))}  exclusions: {len(lists.get('exclusions', []))}")
            print(f"       booking_rules: {len(lists.get('booking_rules', []))}  travel_rules: {len(lists.get('travel_rules', []))}")
            print()
        else:
            exists = db.query(Package).filter(Package.title == title).first()
            if exists:
                for k, v in fields.items():
                    setattr(exists, k, v)
                db.commit()
                print(f"   🔄  Updated  package '{title}' → id={exists.id}")
                updated += 1
            else:
                record = Package(title=title, created_at=now, **fields)
                db.add(record)
                db.commit()
                db.refresh(record)
                print(f"   ✅  Inserted package '{title}' → id={record.id}")
                inserted += 1

    return inserted, updated


def import_categories(categories, db, dry_run):
    inserted = updated = 0

    for cat_data in categories:
        slug = cat_data["slug"]

        if dry_run:
            print(f"   ✅  [DRY RUN] Category '{cat_data['name']}' (slug={slug})")
            print(f"       icon: {cat_data.get('icon') or '—'}")
            print()
            continue

        existing = db.query(PackageCategory).filter(PackageCategory.slug == slug).first()
        if existing:
            existing.name        = cat_data["name"]
            existing.description = cat_data["description"]
            existing.icon        = cat_data.get("icon")
            db.commit()
            print(f"   🔄  Updated category '{cat_data['name']}' (slug={slug})")
            updated += 1
        else:
            record = PackageCategory(
                slug        = slug,
                name        = cat_data["name"],
                description = cat_data["description"],
                icon        = cat_data.get("icon"),
            )
            db.add(record)
            db.commit()
            print(f"   ✅  Inserted category '{cat_data['name']}' (slug={slug})")
            inserted += 1

    return inserted, updated


def import_destinations(destinations, db, dry_run):
    inserted = updated = skipped = 0
    now = str(datetime.utcnow())

    for dest_data in destinations:
        slug = dest_data["slug"]

        if dry_run:
            print(f"   ✅  [DRY RUN] Destination '{dest_data['name']}' (slug={slug})")
            print(f"       image_url: {dest_data.get('image_url') or '—'}")
            print(f"       banner_url: {dest_data.get('banner_url') or '—'}")
            print(f"       is_featured: {dest_data['is_featured']}")
            print()
            continue

        existing = db.query(Destination).filter(Destination.slug == slug).first()
        if existing:
            existing.name        = dest_data["name"]
            existing.description = dest_data["description"]
            existing.image_url   = dest_data.get("image_url")
            existing.banner_url  = dest_data.get("banner_url")
            existing.is_featured = dest_data["is_featured"]
            db.commit()
            print(f"   🔄  Updated destination '{dest_data['name']}' (slug={slug})")
            updated += 1
        else:
            record = Destination(
                slug        = slug,
                name        = dest_data["name"],
                description = dest_data["description"],
                image_url   = dest_data.get("image_url"),
                banner_url  = dest_data.get("banner_url"),
                is_featured = dest_data["is_featured"],
                created_at  = now,
            )
            db.add(record)
            db.commit()
            print(f"   ✅  Inserted destination '{dest_data['name']}' (slug={slug})")
            inserted += 1

    return inserted, updated


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    dry_run = "--dry-run" in sys.argv
    args    = [a for a in sys.argv[1:] if not a.startswith("--")]

    if not args:
        print("Usage: python scripts/import_packages.py <excel_file_or_google_sheets_url> [--dry-run]")
        sys.exit(1)

    wb = load_workbook(args[0])
    print()

    if "Packages" not in wb.sheetnames:
        print("❌  'Packages' sheet not found. Aborting.")
        sys.exit(1)

    # ── Read all sheets ──
    packages     = read_packages(wb["Packages"])
    itinerary    = read_itinerary(wb["Itinerary"])      if "Itinerary"    in wb.sheetnames else {}
    hotels       = read_hotels(wb["Hotels"])             if "Hotels"       in wb.sheetnames else {}
    transport    = read_transport(wb["Transport"])       if "Transport"    in wb.sheetnames else {}
    addons       = read_addons(wb["Addons"])             if "Addons"       in wb.sheetnames else {}
    lists_data   = read_lists(wb["Lists"])               if "Lists"        in wb.sheetnames else {}
    destinations = read_destinations(wb["Destinations"]) if "Destinations" in wb.sheetnames else []
    categories   = read_categories(wb["Categories"])     if "Categories"   in wb.sheetnames else []

    print(f"📦  Packages found    : {len(packages)}")
    print(f"🌍  Destinations found: {len(destinations)}")
    print(f"🗂️   Categories found  : {len(categories)}\n")

    # ── Validate ──
    errors = validate_packages(packages) + validate_destinations(destinations)
    if errors:
        print("❌  Validation errors:\n")
        for e in errors:
            print(f"   • {e}")
        sys.exit(1)

    if dry_run:
        print("🔍  DRY RUN — nothing will be written to the database.\n")

    db = None if dry_run else SessionLocal()

    try:
        # ── Import categories first (packages reference them) ──
        if categories:
            print("── Categories ────────────────────────────────")
            cat_ins, cat_upd = import_categories(categories, db, dry_run)

        # ── Import destinations (packages reference them too) ──
        if destinations:
            print("── Destinations ──────────────────────────────")
            dest_ins, dest_upd = import_destinations(destinations, db, dry_run)

        # ── Import packages ──
        if packages:
            print("── Packages ─────────────────────────────────")
            pkg_ins, pkg_skipped = import_packages(
                packages, itinerary, hotels, transport, addons, lists_data, db, dry_run
            )

    except Exception as e:
        if db:
            db.rollback()
        print(f"\n❌  Error during import: {e}")
        raise
    finally:
        if db:
            db.close()

    print()
    if dry_run:
        print("🔍  Dry run complete — nothing was written to the database.")
    else:
        print("── Summary ───────────────────────────────────")
        if categories:
            print(f"   Categories   — {cat_ins} inserted, {cat_upd} updated")
        if destinations:
            print(f"   Destinations — {dest_ins} inserted, {dest_upd} updated")
        if packages:
            print(f"   Packages    — {pkg_ins} inserted, {pkg_skipped} skipped")
        print("\n🎉  Import complete.")


if __name__ == "__main__":
    main()
